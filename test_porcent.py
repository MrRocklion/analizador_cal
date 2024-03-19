import pandas as pd
import math
from openpyxl import load_workbook
def calcularPorcent(data,porcent,final_path,condition,cal):
    df = data
    longitud = len(df)
    aux_porcent = porcent
    muestra_generada = []
    if condition == 1:
        df_ordenado = df.sample(frac=aux_porcent).sort_values(by=['year', 'mes'], ascending=[True, True])
        muestra_generada = df_ordenado.to_dict(orient='records')
    elif condition == 2:
        N_poblation = longitud
        error = 0.05 
        confianza = 1.96 
        total_porcent = (N_poblation *(confianza**2)*0.5*0.5)/((error**2)*(N_poblation-1)+(confianza**2)*0.5*0.5)
        df_ordenado = df.sample(int(math.floor(total_porcent))).sort_values(by=['year', 'mes'], ascending=[True, True])
        muestra_generada = df_ordenado.to_dict(orient='records')
    else:
        df_ordenado = df.sort_values(by=['year', 'mes'], ascending=[True, True])
        muestra_generada = df_ordenado.to_dict(orient='records')

    if cal == 20:
        workbook = load_workbook('./plantillas/cal20_auditar.xlsx')
        sheet = workbook.active
        aux_init = 3
        
        for i in muestra_generada:
            sheet[f'A{aux_init}'] = i['year']
            sheet[f'B{aux_init}'] = i['mes']
            sheet[f'C{aux_init}'] = i['dia']
            sheet[f'D{aux_init}'] = i['subestacion']
            sheet[f'E{aux_init}'] = i['geo_x']
            sheet[f'F{aux_init}'] = i['geo_y']
            sheet[f'G{aux_init}'] = i['provincia']
            sheet[f'H{aux_init}'] = i['canton']
            sheet[f'K{aux_init}'] = i['registros']
            sheet[f'L{aux_init}'] = i['fase_av']
            sheet[f'M{aux_init}'] = i['fase_bv']
            sheet[f'N{aux_init}'] = i['fase_cv']
            sheet[f'T{aux_init}'] = i['observaciones']
            aux_init += 1
        
        workbook.save(final_path)
        workbook.close()
    elif cal == 30:
        workbook = load_workbook('./plantillas/cal30_auditar.xlsx')
        sheet = workbook.active
        aux_init = 4
        for i in muestra_generada:
            sheet[f'A{aux_init}'] = i['year']
            sheet[f'B{aux_init}'] = i['mes']
            sheet[f'C{aux_init}'] = i['dia']
            sheet[f'D{aux_init}'] = i['codigo']
            sheet[f'E{aux_init}'] = i['tipo']
            sheet[f'F{aux_init}'] = i['subestacion']
            sheet[f'G{aux_init}'] = i['alimentador']
            sheet[f'H{aux_init}'] = i['fases']
            sheet[f'I{aux_init}'] = i['ff']
            sheet[f'J{aux_init}'] = i['fn']
            sheet[f'L{aux_init}'] = i['registros']
            sheet[f'M{aux_init}'] = i['fase_av']
            sheet[f'N{aux_init}'] = i['fase_apst']
            sheet[f'O{aux_init}'] = i['fase_avthd']
            sheet[f'P{aux_init}'] = i['fase_cv']
            sheet[f'Q{aux_init}'] = i['fase_cpst']
            sheet[f'R{aux_init}'] = i['fase_cvthd']
            sheet[f'S{aux_init}'] = i['fase_bv']
            sheet[f'T{aux_init}'] = i['fase_bpst']
            sheet[f'U{aux_init}'] = i['fase_bvthd']
            sheet[f'V{aux_init}'] = i['desequilibrio']
            sheet[f'AI{aux_init}'] = i['observaciones']
            aux_init += 1
        workbook.save(final_path)
        workbook.close()
    elif cal == 40:
        workbook = load_workbook('./plantillas/cal40_auditar.xlsx')
        sheet = workbook.active
        aux_init = 3
        for i in muestra_generada:
            sheet[f'A{aux_init}'] = i['year']
            sheet[f'B{aux_init}'] = i['mes']
            sheet[f'C{aux_init}'] = i['dia']
            sheet[f'D{aux_init}'] = i['codigo']
            sheet[f'E{aux_init}'] = i['tipo']
            sheet[f'F{aux_init}'] = i['provincia']
            sheet[f'G{aux_init}'] = i['canton']
            sheet[f'H{aux_init}'] = i['subestacion']
            sheet[f'I{aux_init}'] = i['alimentador']
            sheet[f'J{aux_init}'] = i['transformador']
            sheet[f'K{aux_init}'] = i['fases']
            sheet[f'L{aux_init}'] = i['ff']
            sheet[f'M{aux_init}'] = i['fn']
            sheet[f'O{aux_init}'] = i['registros']
            sheet[f'P{aux_init}'] = i['fase_av']
            sheet[f'Q{aux_init}'] = i['fase_bv']
            sheet[f'R{aux_init}'] = i['fase_cv']
            sheet[f'W{aux_init}'] = i['observaciones']
        
            aux_init += 1
        workbook.save(final_path)
        workbook.close()
    elif cal == 50:
        workbook = load_workbook('./plantillas/cal50_auditar.xlsx')
        sheet = workbook.active
        aux_init = 4
        for i in muestra_generada:
            sheet[f'A{aux_init}'] = i['year']
            sheet[f'B{aux_init}'] = i['mes']
            sheet[f'C{aux_init}'] = i['dia']
            sheet[f'D{aux_init}'] = i['codigo']
            sheet[f'E{aux_init}'] = i['provincia']
            sheet[f'F{aux_init}'] = i['canton']
            sheet[f'G{aux_init}'] = i['subestacion']
            sheet[f'H{aux_init}'] = i['alimentador']
            sheet[f'I{aux_init}'] = i['ff']
            sheet[f'J{aux_init}'] = i['fn']
            sheet[f'L{aux_init}'] = i['registros']
            sheet[f'M{aux_init}'] = i['fase_av']
            sheet[f'N{aux_init}'] = i['fase_apst']
            sheet[f'O{aux_init}'] = i['fase_avthd']
            sheet[f'P{aux_init}'] = i['fase_bv']
            sheet[f'Q{aux_init}'] = i['fase_bpst']
            sheet[f'R{aux_init}'] = i['fase_bvthd']
            sheet[f'S{aux_init}'] = i['fase_cv']
            sheet[f'T{aux_init}'] = i['fase_cpst']
            sheet[f'U{aux_init}'] = i['fase_cvthd']
            sheet[f'V{aux_init}'] = i['desequilibrio']
            sheet[f'AI{aux_init}'] = i['observaciones']

            aux_init += 1
        workbook.save(final_path)
        workbook.close()

        
data = pd.read_csv('020.csv')
calcularPorcent(data=data,porcent=0.2,final_path='test.xlsx',condition=1,cal=20)