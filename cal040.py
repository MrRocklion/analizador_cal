import openpyxl
import os
from datetime import datetime
from openpyxl import load_workbook
#funcion para ordenar diccionarios
def ordenar_diccionario(total_data):
    for data in total_data:
        for clave, valor in datos_meses.items():
            if data['mes'] == clave:
                data['mes'] = valor
    ordenado_por_dia = sorted(total_data, key=lambda x: x['fila'])         
    ordenado_por_mes = sorted(ordenado_por_dia, key=lambda x: x['mes'])
    ordenado_por_year = sorted(ordenado_por_mes, key=lambda x: x['year'])
    return ordenado_por_year

#
datos_meses = {'Enero':1,'Febrero':2,'Marzo':3,'Abril':4,'Mayo':5,'Junio':6,'Julio':7,'Agosto':8,'Septiembre':9,'Octubre':10,'Noviembre':11,'Diciembre':12}
encabezados = {
                'A': 'fila','B': 'codigo', 'C':'tipo', 'D':'geo_x', 'E':'geo_y','F':'geo_z', 'G': 'provincia',
                'H': 'canton', 'I': 'subestacion', 'J': 'alimentador', 'K': 'transformador', 'L': 'fases',
                'M': 'ff', 'N': 'fn', 'O': 'fecha_inicio', 'P': 'hora_inicio', 'Q': 'fecha_final',
                'R': 'hora_final', 'S': 'registros', 'T': 'fase_av', 'W': 'fase_bv', 'Z': 'fase_cv', 'AG': 'fa8dv9',
                'AH': 'fa9dv10', 'AI': 'fa10dv11', 'AJ': 'fa11dv12', 'AK':'fa12dv13', 'AL':'fa13dv14', 'AM':'fa14dv15', 
                'AN': 'fa15dv','AO':'total_a', 'AS':'fb8dv9', 'AT':'fb9dv10', 'AU':'fb10dv11', 'AV':'fb11dv12', 
                'AW': 'fb12dv13', 'AX':'fb13dv14', 'AY':'fb14dv15', 'AZ':'fb15dv','BA':'total_b', 'BE':'fc8dv9', 
                'BF': 'fc9dv10', 'BG': 'fc10dv11', 'BH':'fc11dv12', 'BI': 'fc12dv13', 'BJ': 'fc13dv14', 
                'BK': 'fc14dv15', 'BL': 'fc15dv','BM':'total_c','BN':'observaciones'
                }

#esta variable contendra la base de datos
data_base = []
def process_sheet(path,file_dir):
    workbook = openpyxl.load_workbook(path+'/'+file_dir,data_only=True)
    lista_de_hojas = workbook.sheetnames
    #esta variable alamacera toda la informacion recopilada de los excels
    data_captured = []
    #abrimos el excel
    #primer for para seleccionar la hoja de calculo que contenga la informacion
    target = ''
    for target_name in lista_de_hojas:
        if target_name[0:3] == 'CAL':
            target = target_name
            break

    # una vez que tenemos el nombre de la hoja que contiene los datos procedemos a abrir esa hoja
    #la variable woorbook contiene el excel con los datos
    sheet_target = workbook[target]
    start_row = 12

    for fila in range(start_row,100):
        if sheet_target[f'B{fila}'].value == None:
            break
        dict_aux = {}
        for clave, valor in encabezados.items():
            cell_name = f'{clave}{fila}'
            if cell_name == f'T{fila}':
                aux_data = round(sheet_target[cell_name].value*100)
                dict_aux[valor] = aux_data
            elif cell_name == f'W{fila}':
                aux_data = round(sheet_target[cell_name].value*100)
                dict_aux[valor] = aux_data
            elif cell_name == f'Z{fila}':
                aux_data = round(sheet_target[cell_name].value*100)
                dict_aux[valor] = aux_data
            else:
                current_cell = sheet_target[cell_name].value
                dict_aux[valor] = current_cell
        empty_dict = {}
        empty_dict['year'] = sheet_target['D3'].value
        fecha  = sheet_target['D4'].value
        fecha_aux = datetime.strptime(str(fecha), "%Y-%m-%d %H:%M:%S")
        empty_dict['mes'] = fecha_aux.month
       
        ##  empieza formato de la fecha
        fecha_aux = str(dict_aux['fecha_inicio'])
        fecha_formated = '0-0-0'
        if len(fecha_aux) > 10:
            year = fecha_aux[0:4]
            mes = fecha_aux[5:7]
            dia = fecha_aux[8:10]
            fecha_formated = f'{dia}-{mes}-{year}'
        else:
            dia = fecha_aux[0:2]
            mes = fecha_aux[3:5]
            year = fecha_aux[6:]
            fecha_formated = f'{dia}-{mes}-{year}'
        dict_aux['fecha_inicio'] = fecha_formated
        empty_dict['dia'] = fecha_formated[0:2]
        empty_dict['file'] = file_dir
        fecha_aux = str(dict_aux['fecha_final'])
        fecha_formated = '0-0-0'
        if len(fecha_aux) > 10:
            year = fecha_aux[0:4]
            mes = fecha_aux[5:7]
            dia = fecha_aux[8:10]
            fecha_formated = f'{dia}-{mes}-{year}'
        else:
            dia = fecha_aux[0:2]
            mes = fecha_aux[3:5]
            year = fecha_aux[6:]
            fecha_formated = f'{dia}-{mes}-{year}'
        dict_aux['fecha_final'] = fecha_formated
        ## desde aqui se deja de dar formato a la fecha
        empty_dict.update(dict_aux)
        data_captured.append(empty_dict)
    return data_captured

def calcular040(path_source,path_final):
    # primero revisamos la cantidad de excels que estan en el directorio actual
    # Obtiene el directorio actual
    # Ruta del directorio que quieres listar
    directorio = path_source

    # Obtener la lista de archivos en el directorio
    archivos = os.listdir(directorio)

    # Imprime la lista de archivos
    listado_archivos = []
    for archivo in archivos:
        if archivo[-5:] == '.xlsx':
            listado_archivos.append(archivo)


    for path_target in listado_archivos:
        data_base.extend(process_sheet(directorio,path_target))
    data_ordenada = ordenar_diccionario(data_base)
    workbook = load_workbook('./plantillas/cal40_consolidado.xlsx')
    sheet = workbook.active
    aux_init = 6
    for i in data_ordenada:
        sheet[f'A{aux_init}'] = i['year']
        sheet[f'B{aux_init}'] = i['mes']
        sheet[f'C{aux_init}'] = i['fila']
        sheet[f'D{aux_init}'] = i['codigo']
        sheet[f'E{aux_init}'] = i['tipo']
        sheet[f'F{aux_init}'] = i['geo_x']
        sheet[f'G{aux_init}'] = i['geo_y']
        sheet[f'H{aux_init}'] = i['geo_z']
        sheet[f'I{aux_init}'] = i['provincia']
        sheet[f'J{aux_init}'] = i['canton']
        sheet[f'K{aux_init}'] = i['subestacion']
        sheet[f'L{aux_init}'] = i['alimentador']
        sheet[f'M{aux_init}'] = i['transformador']
        sheet[f'N{aux_init}'] = i['fases']
        sheet[f'O{aux_init}'] = i['ff']
        sheet[f'P{aux_init}'] = i['fn']
        sheet[f'Q{aux_init}'] = i['fecha_inicio']
        sheet[f'R{aux_init}'] = i['hora_inicio']
        sheet[f'S{aux_init}'] = i['fecha_final']
        sheet[f'T{aux_init}'] = i['hora_final']
        sheet[f'U{aux_init}'] = i['registros']
        sheet[f'V{aux_init}'] = i['fase_av']
        sheet[f'Y{aux_init}'] = i['fase_bv']
        sheet[f'AB{aux_init}'] = i['fase_cv']
        #FASE A
        sheet[f'AI{aux_init}'] = i['fa8dv9']
        sheet[f'AJ{aux_init}'] = i['fa9dv10']
        sheet[f'AK{aux_init}'] = i['fa10dv11']
        sheet[f'AL{aux_init}'] = i['fa11dv12']
        sheet[f'AM{aux_init}'] = i['fa12dv13']
        sheet[f'AN{aux_init}'] = i['fa13dv14']
        sheet[f'AO{aux_init}'] = i['fa14dv15']
        sheet[f'AP{aux_init}'] = i['fa15dv']
        sheet[f'AQ{aux_init}'] = i['total_a']
        #FASE B
        sheet[f'AU{aux_init}'] = i['fb8dv9']
        sheet[f'AV{aux_init}'] = i['fb9dv10']
        sheet[f'AW{aux_init}'] = i['fb10dv11']
        sheet[f'AX{aux_init}'] = i['fb11dv12']
        sheet[f'AY{aux_init}'] = i['fb12dv13']
        sheet[f'AZ{aux_init}'] = i['fb13dv14']
        sheet[f'BA{aux_init}'] = i['fb14dv15']
        sheet[f'BB{aux_init}'] = i['fb15dv']
        sheet[f'BC{aux_init}'] = i['total_b']
        #FACE C
        sheet[f'BG{aux_init}'] = i['fc8dv9']
        sheet[f'BH{aux_init}'] = i['fc9dv10']
        sheet[f'BI{aux_init}'] = i['fc10dv11']
        sheet[f'BJ{aux_init}'] = i['fc11dv12']
        sheet[f'BK{aux_init}'] = i['fc12dv13']
        sheet[f'BL{aux_init}'] = i['fc13dv14']
        sheet[f'BM{aux_init}'] = i['fc14dv15']
        sheet[f'BN{aux_init}'] = i['fc15dv']
        sheet[f'BO{aux_init}'] = i['total_c']
        sheet[f'BP{aux_init}'] = i['observaciones']
        #FILE
        sheet[f'BQ{aux_init}']=i['file']
        aux_init += 1
    workbook.save(path_final)
    workbook.close()
    return data_ordenada