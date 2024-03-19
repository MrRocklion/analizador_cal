import openpyxl
import os
from openpyxl import load_workbook
from datetime import datetime

#funcion para ordenar diccionarios
datos_meses = {'Enero':1,'Febrero':2,'Marzo':3,'Abril':4,'Mayo':5,'Junio':6,'Julio':7,'Agosto':8,'Septiembre':9,'Octubre':10,'Noviembre':11,'Diciembre':12}
def ordenar_diccionario(total_data):
    for data in total_data:
        for clave, valor in datos_meses.items():
            if data['mes'] == clave:
                data['mes'] = valor
    ordenado_por_dia = sorted(total_data, key=lambda x: x['fila'])         
    ordenado_por_mes = sorted(ordenado_por_dia, key=lambda x: x['mes'])
    ordenado_por_year = sorted(ordenado_por_mes, key=lambda x: x['year'])
    return ordenado_por_year
##### 


def process_sheet(path,file_dir):
    encabezados = { 
                'A': 'fila','B':'codigo', 'C': 'tipo', 'D': 'geo_x', 'E': 'geo_y','F': 'geo_z', 'G': 'provincia', 'H': 'canton', 'I': 'subestacion',
                'J': 'alimentador', 'L': 'fases', 'M': 'ff', 'N': 'fn', 'O': 'fecha_inicio', 'P': 'hora_inicio',
                'Q': 'fecha_final', 'R': 'hora_final', 'S': 'registros', 'T': 'fase_av', 'U': 'fase_apst', 'V': 'fase_avthd',
                'W': 'fase_bv', 'X': 'fase_bpst','Y': 'fase_bvthd', 'Z': 'fase_cv', 'AA': 'fase_cpst', 'AB': 'fase_cvthd', 'AC': 'desequilibrio',
                'AG': 'fa8dv9','AH': 'fa9dv10', 'AI': 'fa10dv11', 'AJ': 'fa11dv12', 'AK': 'fa12dv13', 'AL': 'fa13dv14',
                'AM': 'fa14dv15','AN': 'fa15dv','AO':'total_a', 'AS': 'fb8dv9', 'AT':'fb9dv10', 'AU':'fb10dv11', 'AV': 'fb11dv12',
                'AW': 'fb12dv13','AX': 'fb13dv14', 'AY': 'fb14dv15', 'AZ': 'fb15dv', 'BA': 'total_b','BE':'fc8dv9', 'BF':'fc9dv10',
                'BG':'fc10dv11', 'BH': 'fc11dv12', 'BI': 'fc12dv13', 'BJ':'fc13dv14', 'BK': 'fc14dv15', 'BL': 'fc15dv','BM': 'total_c','BN': 'observaciones', 
                'BR':'fa_armonico2','BS':'fb_armonico2','BT':'fc_armonico2',
                'BU':'fa_armonico3','BV':'fb_armonico3','BW':'fc_armonico3',
                'BX':'fa_armonico4','BY':'fb_armonico4','BZ':'fc_armonico4',
                'CA':'fa_armonico5','CB':'fb_armonico5','CC':'fc_armonico5',
                'CD':'fa_armonico6','CE':'fb_armonico6','CF':'fc_armonico6',
                'CG':'fa_armonico7','CH':'fb_armonico7','CI':'fc_armonico7',
                'CJ':'fa_armonico8','CK':'fb_armonico8','CL':'fc_armonico8',
                'CM':'fa_armonico9','CN':'fb_armonico9','CO':'fc_armonico9',
                'CP':'fa_armonico10','CQ':'fb_armonico10','CR':'fc_armonico10',
                'CS':'fa_armonico11','CT':'fb_armonico11','CU':'fc_armonico11',
                'CV':'fa_armonico12','CW':'fb_armonico12','CX':'fc_armonico12',
                'CY':'fa_armonico13','CZ':'fb_armonico13','DA':'fc_armonico13',
                'DB':'fa_armonico14','DC':'fb_armonico14','DD':'fc_armonico14',
                'DE':'fa_armonico15','DF':'fb_armonico15','DG':'fc_armonico15',
                'DH':'fa_armonico16','DI':'fb_armonico16','DJ':'fc_armonico16',
                'DK':'fa_armonico17','DL':'fb_armonico17','DM':'fc_armonico17',
                'DN':'fa_armonico18','DO':'fb_armonico18','DP':'fc_armonico18',
                'DQ':'fa_armonico19','DR':'fb_armonico19','DS':'fc_armonico19',
                'DT':'fa_armonico20','DU':'fb_armonico20','DV':'fc_armonico20',
                'DW':'fa_armonico21','DX':'fb_armonico21','DY':'fc_armonico21',
                'DZ':'fa_armonico22','EA':'fb_armonico22','EB':'fc_armonico22',
                'EC':'fa_armonico23','ED':'fb_armonico23','EE':'fc_armonico23',
                'EF':'fa_armonico24','EG':'fb_armonico24','EH':'fc_armonico24',
                'EI':'fa_armonico25','EJ':'fb_armonico25','EK':'fc_armonico25',
                'EL':'fa_armonico26','EM':'fb_armonico26','EN':'fc_armonico26',
                'EO':'fa_armonico27','EP':'fb_armonico27','EQ':'fc_armonico27',
                'ER':'fa_armonico28','ES':'fb_armonico28','ET':'fc_armonico28',
                'EU':'fa_armonico29','EV':'fb_armonico29','EW':'fc_armonico29',
                'EX':'fa_armonico30','EY':'fb_armonico30','EZ':'fc_armonico30',
                'FA':'fa_armonico31','FB':'fb_armonico31','FC':'fc_armonico31',
                'FD':'fa_armonico32','FE':'fb_armonico32','FF':'fc_armonico32',
                'FG':'fa_armonico33','FH':'fb_armonico33','FI':'fc_armonico33',
                'FJ':'fa_armonico34','FK':'fb_armonico34','FL':'fc_armonico34',
                'FM':'fa_armonico35','FN':'fb_armonico35','FO':'fc_armonico35',
                'FP':'fa_armonico36','FQ':'fb_armonico36','FR':'fc_armonico36',
                'FS':'fa_armonico37','FT':'fb_armonico37','FU':'fc_armonico37',
                'FV':'fa_armonico38','FW':'fb_armonico38','FX':'fc_armonico38',
                'FY':'fa_armonico39','FZ':'fb_armonico39','GA':'fc_armonico39',
                'GB':'fa_armonico40','GC':'fb_armonico40','GD':'fc_armonico40',
                'GE':'fa_armonico41','GF':'fb_armonico41','GG':'fc_armonico41',
                'GH':'fa_armonico42','GI':'fb_armonico42','GJ':'fc_armonico42',
                'GK':'fa_armonico43','GL':'fb_armonico43','GM':'fc_armonico43',
                'GN':'fa_armonico44','GO':'fb_armonico44','GP':'fc_armonico44',
                'GQ':'fa_armonico45','GR':'fb_armonico45','GS':'fc_armonico45',
                'GT':'fa_armonico46','GU':'fb_armonico46','GV':'fc_armonico46',
                'GW':'fa_armonico47','GX':'fb_armonico47','GY':'fc_armonico47',
                'GZ':'fa_armonico48','HA':'fb_armonico48','HB':'fc_armonico48',
                'HC':'fa_armonico49','HD':'fb_armonico49','HE':'fc_armonico49',
                'HF':'fa_armonico50','HG':'fb_armonico50','HH':'fc_armonico50'
                }

    workbook = openpyxl.load_workbook(path+'/'+file_dir,data_only=True)
    lista_de_hojas = workbook.sheetnames
    #esta variable alamacera toda la informacion recopilada de los excels
    data_captured = []
    #abrimos el excel
    #primer for para seleccionar la hoja de calculo que contenga la informacion
    target = ''
    for target_name in lista_de_hojas:
        if target_name[0:3] == 'CAL':
            target = target_name
            break

    # una vez que tenemos el nombre de la hoja que contiene los datos procedemos a abrir esa hoja
    #la variable woorbook contiene el excel con los datos
    sheet_target = workbook[target]
    start_row = 12

    for fila in range(start_row,100):
        if sheet_target[f'B{fila}'].value == None:
            break
        dict_aux = {}
        for clave, valor in encabezados.items():
            cell_name = f'{clave}{fila}'

            if cell_name == f'T{fila}':
                aux_data = round(sheet_target[cell_name].value*100)
                dict_aux[valor] = aux_data
            # elif cell_name == f'BG{fila}':
            #     current_cell = sheet_target[cell_name].value
            #     dict_aux[valor] = current_cell
            #     print(valor)
            #     print(current_cell)
            #     print(file_dir)
            elif cell_name == f'U{fila}':
                aux_data = round(sheet_target[cell_name].value*100,2)
                dict_aux[valor] = aux_data
            elif cell_name == f'V{fila}':
                aux_data = round(sheet_target[cell_name].value*100,2)
                dict_aux[valor] = aux_data
            elif cell_name == f'W{fila}':
                val = sheet_target[cell_name].value
                if val == None:
                    dict_aux[valor] = val
                else:
                    aux_data = round(val*100,2)
                    dict_aux[valor] = aux_data
            elif cell_name == f'X{fila}':
                val = sheet_target[cell_name].value
                if val == None:
                    dict_aux[valor] = val
                else:
                    aux_data = round(val*100,2)
                    dict_aux[valor] = aux_data
            elif cell_name == f'Y{fila}':
                val = sheet_target[cell_name].value
                if val == None:
                    dict_aux[valor] = val
                else:
                    aux_data = round(val*100,2)
                    dict_aux[valor] = aux_data
            elif cell_name == f'Z{fila}':
                val = sheet_target[cell_name].value
                if val == None:
                    dict_aux[valor] = val
                else:
                    aux_data = round(val*100,2)
                    dict_aux[valor] = aux_data
            elif cell_name == f'AA{fila}':
                val = sheet_target[cell_name].value
                if val == None:
                    dict_aux[valor] = val
                else:
                    aux_data = round(val*100,2)
                    dict_aux[valor] = aux_data
            elif cell_name == f'AC{fila}':
                if sheet_target[cell_name].value == None:
                    dict_aux[valor] = ''
                else:
                    aux_data = round(sheet_target[cell_name].value*100,2)
                    dict_aux[valor] = aux_data
            else:
                current_cell = sheet_target[cell_name].value
                dict_aux[valor] = current_cell
        empty_dict = {}
        empty_dict['year'] = sheet_target['D3'].value
        fecha  = sheet_target['D4'].value
        fecha_aux = datetime.strptime(str(fecha), "%Y-%m-%d %H:%M:%S")
        empty_dict['mes'] = fecha_aux.month

 
            ##  empieza formato de la fecha
        fecha_aux = str(dict_aux['fecha_inicio'])
        fecha_formated = '0-0-0'
        if len(fecha_aux) > 10:
            year = fecha_aux[0:4]
            mes = fecha_aux[5:7]
            dia = fecha_aux[8:10]
            fecha_formated = f'{dia}-{mes}-{year}'
        else:
            dia = fecha_aux[0:2]
            mes = fecha_aux[3:5]
            year = fecha_aux[6:]
            fecha_formated = f'{dia}-{mes}-{year}'
        dict_aux['fecha_inicio'] = fecha_formated
        empty_dict['dia'] = fecha_formated[0:2]
        empty_dict['file'] = file_dir
        fecha_aux = str(dict_aux['fecha_final'])
        fecha_formated = '0-0-0'
        if len(fecha_aux) > 10:
            year = fecha_aux[0:4]
            mes = fecha_aux[5:7]
            dia = fecha_aux[8:10]
            fecha_formated = f'{dia}-{mes}-{year}'
        else:
            dia = fecha_aux[0:2]
            mes = fecha_aux[3:5]
            year = fecha_aux[6:]
            fecha_formated = f'{dia}-{mes}-{year}'
        dict_aux['fecha_final'] = fecha_formated
        ## desde aqui se deja de dar formato a la fecha
        empty_dict.update(dict_aux)
        data_captured.append(empty_dict)
    return data_captured

def calcular030(path_source,path_final):
    data_base = []

    directorio = path_source


    archivos = os.listdir(directorio)

    # Imprime la lista de archivos
    listado_archivos = []
    for archivo in archivos:
        if archivo[-5:] == '.xlsx':
            listado_archivos.append(archivo)

    for path_target in listado_archivos:
        data_base.extend(process_sheet(directorio,path_target))

    data_ordenada = ordenar_diccionario(data_base)
    workbook = load_workbook('./plantillas/cal30_consolidado.xlsx')
    sheet = workbook.active
    aux_init = 6
    for i in data_ordenada:
        sheet[f'A{aux_init}'] = i['year']
        sheet[f'B{aux_init}'] = i['mes']
        sheet[f'C{aux_init}'] = i['fila']
        sheet[f'D{aux_init}'] = i['codigo']
        sheet[f'E{aux_init}'] = i['tipo']
        sheet[f'F{aux_init}'] = i['geo_x']
        sheet[f'G{aux_init}'] = i['geo_y']
        sheet[f'H{aux_init}'] = i['geo_z']
        sheet[f'I{aux_init}'] = i['provincia']
        sheet[f'J{aux_init}'] = i['canton']
        sheet[f'K{aux_init}'] = i['subestacion']
        sheet[f'L{aux_init}'] = i['alimentador']
        sheet[f'M{aux_init}'] = i['fases']
        sheet[f'N{aux_init}'] = i['ff']
        sheet[f'O{aux_init}'] = i['fn']
        sheet[f'P{aux_init}'] = i['fecha_inicio']
        sheet[f'Q{aux_init}'] = i['hora_inicio']
        sheet[f'R{aux_init}'] = i['fecha_final']
        sheet[f'S{aux_init}'] = i['hora_final']
        sheet[f'T{aux_init}'] = i['registros']
        #FUERA DE LIMITES
        sheet[f'U{aux_init}'] = i['fase_av']
        sheet[f'V{aux_init}'] = i['fase_apst']
        sheet[f'W{aux_init}'] = i['fase_avthd']
        sheet[f'X{aux_init}'] = i['fase_bv']
        sheet[f'Y{aux_init}'] = i['fase_bpst']
        sheet[f'Z{aux_init}'] = i['fase_bvthd']
        sheet[f'AA{aux_init}'] = i['fase_cv']
        sheet[f'AB{aux_init}'] = i['fase_cpst']
        sheet[f'AC{aux_init}'] = i['fase_cvthd']
        sheet[f'AD{aux_init}'] = i['desequilibrio']
        #FASE A
        sheet[f'AH{aux_init}'] = i['fa8dv9']
        sheet[f'AI{aux_init}'] = i['fa9dv10']
        sheet[f'AJ{aux_init}'] = i['fa10dv11']
        sheet[f'AK{aux_init}'] = i['fa11dv12']
        sheet[f'AL{aux_init}'] = i['fa12dv13']
        sheet[f'AM{aux_init}'] = i['fa13dv14']
        sheet[f'AN{aux_init}'] = i['fa14dv15']
        sheet[f'AO{aux_init}'] = i['fa15dv']
        sheet[f'AP{aux_init}'] = i['total_a']
        #FASE B
        sheet[f'AT{aux_init}'] = i['fb8dv9']
        sheet[f'AU{aux_init}'] = i['fb9dv10']
        sheet[f'AV{aux_init}'] = i['fb10dv11']
        sheet[f'AW{aux_init}'] = i['fb11dv12']
        sheet[f'AX{aux_init}'] = i['fb12dv13']
        sheet[f'AY{aux_init}'] = i['fb13dv14']
        sheet[f'AZ{aux_init}'] = i['fb14dv15']
        sheet[f'BA{aux_init}'] = i['fb15dv']
        sheet[f'BB{aux_init}'] = i['total_b']
        #FACE C
        sheet[f'BF{aux_init}'] = i['fc8dv9']
        sheet[f'BG{aux_init}'] = i['fc9dv10']
        sheet[f'BH{aux_init}'] = i['fc10dv11']
        sheet[f'BI{aux_init}'] = i['fc11dv12']
        sheet[f'BJ{aux_init}'] = i['fc12dv13']
        sheet[f'BK{aux_init}'] = i['fc13dv14']
        sheet[f'BL{aux_init}'] = i['fc14dv15']
        sheet[f'BM{aux_init}'] = i['fc15dv']
        sheet[f'BN{aux_init}'] = i['total_c']
        sheet[f'BO{aux_init}'] = i['observaciones']
        #ARMONICOS FASE A
        sheet[f'BR{aux_init}']=i['fa_armonico2']
        sheet[f'BU{aux_init}']=i['fa_armonico3']
        sheet[f'BX{aux_init}']=i['fa_armonico4']
        sheet[f'CA{aux_init}']=i['fa_armonico5']
        sheet[f'CD{aux_init}']=i['fa_armonico6']
        sheet[f'CG{aux_init}']=i['fa_armonico7']
        sheet[f'CJ{aux_init}']=i['fa_armonico8']
        sheet[f'CM{aux_init}']=i['fa_armonico9']
        sheet[f'CP{aux_init}']=i['fa_armonico10']
        sheet[f'CS{aux_init}']=i['fa_armonico11']
        sheet[f'CV{aux_init}']=i['fa_armonico12']
        sheet[f'CY{aux_init}']=i['fa_armonico13']
        sheet[f'DB{aux_init}']=i['fa_armonico14']
        sheet[f'DE{aux_init}']=i['fa_armonico15']
        sheet[f'DH{aux_init}']=i['fa_armonico16']
        sheet[f'DK{aux_init}']=i['fa_armonico17']
        sheet[f'DN{aux_init}']=i['fa_armonico18']
        sheet[f'DQ{aux_init}']=i['fa_armonico19']
        sheet[f'DT{aux_init}']=i['fa_armonico20']
        sheet[f'DW{aux_init}']=i['fa_armonico21']
        sheet[f'DZ{aux_init}']=i['fa_armonico22']
        sheet[f'EC{aux_init}']=i['fa_armonico23']
        sheet[f'EF{aux_init}']=i['fa_armonico24']
        sheet[f'EI{aux_init}']=i['fa_armonico25']
        sheet[f'EL{aux_init}']=i['fa_armonico26']
        sheet[f'EO{aux_init}']=i['fa_armonico27']
        sheet[f'ER{aux_init}']=i['fa_armonico28']
        sheet[f'EU{aux_init}']=i['fa_armonico29']
        sheet[f'EX{aux_init}']=i['fa_armonico30']
        sheet[f'FA{aux_init}']=i['fa_armonico31']
        sheet[f'FD{aux_init}']=i['fa_armonico32']
        sheet[f'FG{aux_init}']=i['fa_armonico33']
        sheet[f'FJ{aux_init}']=i['fa_armonico34']
        sheet[f'FM{aux_init}']=i['fa_armonico35']
        sheet[f'FP{aux_init}']=i['fa_armonico36']
        sheet[f'FS{aux_init}']=i['fa_armonico37']
        sheet[f'FV{aux_init}']=i['fa_armonico38']
        sheet[f'FY{aux_init}']=i['fa_armonico39']
        sheet[f'GB{aux_init}']=i['fa_armonico40']
        sheet[f'GE{aux_init}']=i['fa_armonico41']
        sheet[f'GH{aux_init}']=i['fa_armonico42']
        sheet[f'GK{aux_init}']=i['fa_armonico43']
        sheet[f'GN{aux_init}']=i['fa_armonico44']
        sheet[f'GQ{aux_init}']=i['fa_armonico45']
        sheet[f'GT{aux_init}']=i['fa_armonico46']
        sheet[f'GW{aux_init}']=i['fa_armonico47']
        sheet[f'GZ{aux_init}']=i['fa_armonico48']
        sheet[f'HC{aux_init}']=i['fa_armonico49']
        sheet[f'HF{aux_init}']=i['fa_armonico50']

        #ARMONICOS FASE B
        sheet[f'BS{aux_init}']=i['fb_armonico2']
        sheet[f'BV{aux_init}']=i['fb_armonico3']
        sheet[f'BY{aux_init}']=i['fb_armonico4']
        sheet[f'CB{aux_init}']=i['fb_armonico5']
        sheet[f'CE{aux_init}']=i['fb_armonico6']
        sheet[f'CH{aux_init}']=i['fb_armonico7']
        sheet[f'CK{aux_init}']=i['fb_armonico8']
        sheet[f'CN{aux_init}']=i['fb_armonico9']
        sheet[f'CQ{aux_init}']=i['fb_armonico10']
        sheet[f'CT{aux_init}']=i['fb_armonico11']
        sheet[f'CW{aux_init}']=i['fb_armonico12']
        sheet[f'CZ{aux_init}']=i['fb_armonico13']
        sheet[f'DC{aux_init}']=i['fb_armonico14']
        sheet[f'DF{aux_init}']=i['fb_armonico15']
        sheet[f'DI{aux_init}']=i['fb_armonico16']
        sheet[f'DL{aux_init}']=i['fb_armonico17']
        sheet[f'DO{aux_init}']=i['fb_armonico18']
        sheet[f'DR{aux_init}']=i['fb_armonico19']
        sheet[f'DU{aux_init}']=i['fb_armonico20']
        sheet[f'DX{aux_init}']=i['fb_armonico21']
        sheet[f'EA{aux_init}']=i['fb_armonico22']
        sheet[f'ED{aux_init}']=i['fb_armonico23']
        sheet[f'EG{aux_init}']=i['fb_armonico24']
        sheet[f'EJ{aux_init}']=i['fb_armonico25']
        sheet[f'EM{aux_init}']=i['fb_armonico26']
        sheet[f'EP{aux_init}']=i['fb_armonico27']
        sheet[f'ES{aux_init}']=i['fb_armonico28']
        sheet[f'EV{aux_init}']=i['fb_armonico29']
        sheet[f'EY{aux_init}']=i['fb_armonico30']
        sheet[f'FB{aux_init}']=i['fb_armonico31']
        sheet[f'FE{aux_init}']=i['fb_armonico32']
        sheet[f'FH{aux_init}']=i['fb_armonico33']
        sheet[f'FK{aux_init}']=i['fb_armonico34']
        sheet[f'FN{aux_init}']=i['fb_armonico35']
        sheet[f'FQ{aux_init}']=i['fb_armonico36']
        sheet[f'FT{aux_init}']=i['fb_armonico37']
        sheet[f'FW{aux_init}']=i['fb_armonico38']
        sheet[f'FZ{aux_init}']=i['fb_armonico39']
        sheet[f'GC{aux_init}']=i['fb_armonico40']
        sheet[f'GF{aux_init}']=i['fb_armonico41']
        sheet[f'GI{aux_init}']=i['fb_armonico42']
        sheet[f'GL{aux_init}']=i['fb_armonico43']
        sheet[f'GO{aux_init}']=i['fb_armonico44']
        sheet[f'GR{aux_init}']=i['fb_armonico45']
        sheet[f'GU{aux_init}']=i['fb_armonico46']
        sheet[f'GX{aux_init}']=i['fb_armonico47']
        sheet[f'HA{aux_init}']=i['fb_armonico48']
        sheet[f'HD{aux_init}']=i['fb_armonico49']
        sheet[f'HG{aux_init}']=i['fb_armonico50']
        #ARMONICOS FASE C
        sheet[f'BT{aux_init}']=i['fc_armonico2']
        sheet[f'BW{aux_init}']=i['fc_armonico3']
        sheet[f'BZ{aux_init}']=i['fc_armonico4']
        sheet[f'CC{aux_init}']=i['fc_armonico5']
        sheet[f'CF{aux_init}']=i['fc_armonico6']
        sheet[f'CI{aux_init}']=i['fc_armonico7']
        sheet[f'CL{aux_init}']=i['fc_armonico8']
        sheet[f'CO{aux_init}']=i['fc_armonico9']
        sheet[f'CR{aux_init}']=i['fc_armonico10']
        sheet[f'CU{aux_init}']=i['fc_armonico11']
        sheet[f'CX{aux_init}']=i['fc_armonico12']
        sheet[f'DA{aux_init}']=i['fc_armonico13']
        sheet[f'DD{aux_init}']=i['fc_armonico14']
        sheet[f'DG{aux_init}']=i['fc_armonico15']
        sheet[f'DJ{aux_init}']=i['fc_armonico16']
        sheet[f'DM{aux_init}']=i['fc_armonico17']
        sheet[f'DP{aux_init}']=i['fc_armonico18']
        sheet[f'DS{aux_init}']=i['fc_armonico19']
        sheet[f'DV{aux_init}']=i['fc_armonico20']
        sheet[f'DY{aux_init}']=i['fc_armonico21']
        sheet[f'EB{aux_init}']=i['fc_armonico22']
        sheet[f'EE{aux_init}']=i['fc_armonico23']
        sheet[f'EH{aux_init}']=i['fc_armonico24']
        sheet[f'EK{aux_init}']=i['fc_armonico25']
        sheet[f'EN{aux_init}']=i['fc_armonico26']
        sheet[f'EQ{aux_init}']=i['fc_armonico27']
        sheet[f'ET{aux_init}']=i['fc_armonico28']
        sheet[f'EW{aux_init}']=i['fc_armonico29']
        sheet[f'EZ{aux_init}']=i['fc_armonico30']
        sheet[f'FC{aux_init}']=i['fc_armonico31']
        sheet[f'FF{aux_init}']=i['fc_armonico32']
        sheet[f'FI{aux_init}']=i['fc_armonico33']
        sheet[f'FL{aux_init}']=i['fc_armonico34']
        sheet[f'FO{aux_init}']=i['fc_armonico35']
        sheet[f'FR{aux_init}']=i['fc_armonico36']
        sheet[f'FU{aux_init}']=i['fc_armonico37']
        sheet[f'FX{aux_init}']=i['fc_armonico38']
        sheet[f'GA{aux_init}']=i['fc_armonico39']
        sheet[f'GD{aux_init}']=i['fc_armonico40']
        sheet[f'GG{aux_init}']=i['fc_armonico41']
        sheet[f'GJ{aux_init}']=i['fc_armonico42']
        sheet[f'GM{aux_init}']=i['fc_armonico43']
        sheet[f'GP{aux_init}']=i['fc_armonico44']
        sheet[f'GS{aux_init}']=i['fc_armonico45']
        sheet[f'GV{aux_init}']=i['fc_armonico46']
        sheet[f'GY{aux_init}']=i['fc_armonico47']
        sheet[f'HB{aux_init}']=i['fc_armonico48']
        sheet[f'HE{aux_init}']=i['fc_armonico49']
        sheet[f'HH{aux_init}']=i['fc_armonico50']
        #FILE
        sheet[f'HI{aux_init}']=i['file']
        aux_init += 1

    workbook.save(path_final)
    workbook.close()
    return data_ordenada




